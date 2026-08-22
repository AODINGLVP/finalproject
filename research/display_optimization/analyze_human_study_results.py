"""Analyze the pre-registered 241-node behavior-tree usability study.

The script never generates participant observations. With an empty workbook it
writes a clearly marked pending report. Once real observations are entered, it
validates consent and scoring fields, creates descriptive CSV data, and performs
the pre-registered repeated-measures analysis without optional statistics
packages. Run ``--self-test`` to validate the mathematics on in-memory synthetic
values; self-test values are never written as study results.
"""

from __future__ import annotations

import argparse
import csv
import itertools
import math
import random
import statistics
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import openpyxl


METHODS = ["Baseline", "Compact", "Collapse", "Fisheye", "Search", "Minimap"]
TASKS = ["Locate Action", "Trace Active Path", "Edit Decorator"]
ALPHA = 0.05
MIN_COMPLETE_PARTICIPANTS = 8
BOOTSTRAP_SAMPLES = 10_000


@dataclass(frozen=True)
class Trial:
    trial_id: str
    participant: str
    task: str
    method: str
    target_key: str
    time_s: float
    errors: float
    zoom: float
    pan: float
    success: bool
    readability: float
    cognitive_load: float
    excluded: bool
    exclusion_reason: str

    @property
    def navigation(self) -> float:
        return self.zoom + self.pan


def _as_float(value: object, label: str, minimum: float, maximum: float) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} must be numeric, received {value!r}") from exc
    if not minimum <= number <= maximum:
        raise ValueError(f"{label}={number} is outside {minimum}..{maximum}")
    return number


def _yes_no(value: object, label: str) -> bool:
    normalized = str(value or "").strip().lower()
    if normalized == "yes":
        return True
    if normalized == "no":
        return False
    raise ValueError(f"{label} must be Yes or No, received {value!r}")


def load_trials(workbook_path: Path) -> tuple[list[Trial], int]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    raw = workbook["Raw Data"]
    participants = workbook["Participants"]
    consent = {
        str(participants.cell(row, 1).value): str(participants.cell(row, 8).value or "").strip()
        for row in range(4, participants.max_row + 1)
        if participants.cell(row, 1).value
    }
    headers = {str(raw.cell(3, column).value): column for column in range(1, raw.max_column + 1)}
    required = [
        "Trial ID", "Participant", "Task", "Method", "Target Key", "Time (s)",
        "Errors", "Zoom", "Pan", "Success", "Readability 1-7",
        "Cognitive Load 1-7", "Excluded", "Exclusion Reason",
    ]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Raw Data is missing required columns: {', '.join(missing)}")
    trials: list[Trial] = []
    planned_rows = 0
    seen_ids: set[str] = set()
    for row in range(4, raw.max_row + 1):
        trial_id = str(raw.cell(row, headers["Trial ID"]).value or "").strip()
        if not trial_id:
            continue
        planned_rows += 1
        time_value = raw.cell(row, headers["Time (s)"]).value
        if time_value in (None, ""):
            continue
        if trial_id in seen_ids:
            raise ValueError(f"duplicate completed Trial ID {trial_id}")
        seen_ids.add(trial_id)
        participant = str(raw.cell(row, headers["Participant"]).value or "").strip()
        if consent.get(participant) != "Yes":
            raise ValueError(f"{trial_id}: participant {participant} does not have Consent=Yes")
        task = str(raw.cell(row, headers["Task"]).value or "").strip()
        method = str(raw.cell(row, headers["Method"]).value or "").strip()
        if task not in TASKS:
            raise ValueError(f"{trial_id}: unknown task {task!r}")
        if method not in METHODS:
            raise ValueError(f"{trial_id}: unknown method {method!r}")
        excluded = _yes_no(raw.cell(row, headers["Excluded"]).value, f"{trial_id} Excluded")
        exclusion_reason = str(raw.cell(row, headers["Exclusion Reason"]).value or "").strip()
        if excluded and not exclusion_reason:
            raise ValueError(f"{trial_id}: excluded trial requires an exclusion reason")
        trials.append(Trial(
            trial_id=trial_id,
            participant=participant,
            task=task,
            method=method,
            target_key=str(raw.cell(row, headers["Target Key"]).value or "").strip(),
            time_s=_as_float(time_value, f"{trial_id} Time", 0.0, 180.0),
            errors=_as_float(raw.cell(row, headers["Errors"]).value, f"{trial_id} Errors", 0.0, 999.0),
            zoom=_as_float(raw.cell(row, headers["Zoom"]).value, f"{trial_id} Zoom", 0.0, 999.0),
            pan=_as_float(raw.cell(row, headers["Pan"]).value, f"{trial_id} Pan", 0.0, 999.0),
            success=_yes_no(raw.cell(row, headers["Success"]).value, f"{trial_id} Success"),
            readability=_as_float(raw.cell(row, headers["Readability 1-7"]).value, f"{trial_id} Readability", 1.0, 7.0),
            cognitive_load=_as_float(raw.cell(row, headers["Cognitive Load 1-7"]).value, f"{trial_id} Cognitive Load", 1.0, 7.0),
            excluded=excluded,
            exclusion_reason=exclusion_reason,
        ))
    return trials, planned_rows


def quartiles(values: Sequence[float]) -> tuple[float, float, float]:
    if not values:
        return math.nan, math.nan, math.nan
    median = statistics.median(values)
    if len(values) == 1:
        return values[0], median, values[0]
    q1, _, q3 = statistics.quantiles(values, n=4, method="inclusive")
    return q1, median, q3


def average_ranks(values: Sequence[float]) -> list[float]:
    order = sorted(range(len(values)), key=lambda index: values[index])
    ranks = [0.0] * len(values)
    cursor = 0
    while cursor < len(order):
        end = cursor + 1
        while end < len(order) and values[order[end]] == values[order[cursor]]:
            end += 1
        average = (cursor + 1 + end) / 2.0
        for position in range(cursor, end):
            ranks[order[position]] = average
        cursor = end
    return ranks


def regularized_gamma_q(a: float, x: float) -> float:
    if x < 0.0 or a <= 0.0:
        return math.nan
    if x == 0.0:
        return 1.0
    epsilon = 3.0e-14
    tiny = 1.0e-300
    max_iterations = 10_000
    if x < a + 1.0:
        term = 1.0 / a
        total = term
        ap = a
        for _ in range(max_iterations):
            ap += 1.0
            term *= x / ap
            total += term
            if abs(term) < abs(total) * epsilon:
                lower = total * math.exp(-x + a * math.log(x) - math.lgamma(a))
                return min(max(1.0 - lower, 0.0), 1.0)
        raise RuntimeError("gamma series did not converge")
    b = x + 1.0 - a
    c = 1.0 / tiny
    d = 1.0 / b
    h = d
    for iteration in range(1, max_iterations + 1):
        an = -float(iteration) * (float(iteration) - a)
        b += 2.0
        d = an * d + b
        if abs(d) < tiny:
            d = tiny
        c = b + an / c
        if abs(c) < tiny:
            c = tiny
        d = 1.0 / d
        delta = d * c
        h *= delta
        if abs(delta - 1.0) < epsilon:
            return min(max(math.exp(-x + a * math.log(x) - math.lgamma(a)) * h, 0.0), 1.0)
    raise RuntimeError("gamma continued fraction did not converge")


def chi_square_survival(statistic: float, degrees_of_freedom: int) -> float:
    if statistic <= 0.0:
        return 1.0
    return regularized_gamma_q(degrees_of_freedom / 2.0, statistic / 2.0)


def friedman_test(matrix: Sequence[Sequence[float]]) -> tuple[float, float]:
    participant_count = len(matrix)
    method_count = len(matrix[0]) if matrix else 0
    if participant_count < 2 or method_count < 2:
        return math.nan, math.nan
    rank_rows = [average_ranks(row) for row in matrix]
    rank_sums = [sum(row[column] for row in rank_rows) for column in range(method_count)]
    statistic = (12.0 / (participant_count * method_count * (method_count + 1.0))) * sum(value * value for value in rank_sums)
    statistic -= 3.0 * participant_count * (method_count + 1.0)
    tie_total = 0.0
    for row in matrix:
        counts: dict[float, int] = defaultdict(int)
        for value in row:
            counts[value] += 1
        tie_total += sum(count ** 3 - count for count in counts.values() if count > 1)
    correction = 1.0 - tie_total / (participant_count * (method_count ** 3 - method_count))
    if correction <= 0.0:
        return 0.0, 1.0
    statistic /= correction
    return statistic, chi_square_survival(statistic, method_count - 1)


def exact_wilcoxon(baseline: Sequence[float], optimized: Sequence[float]) -> tuple[float, float, float, list[float]]:
    differences = [right - left for left, right in zip(baseline, optimized) if right != left]
    if not differences:
        return 0.0, 1.0, 0.0, []
    ranks = average_ranks([abs(value) for value in differences])
    positive = sum(rank for rank, difference in zip(ranks, differences) if difference > 0.0)
    negative = sum(rank for rank, difference in zip(ranks, differences) if difference < 0.0)
    observed = min(positive, negative)
    scaled_ranks = [round(rank * 2.0) for rank in ranks]
    total = sum(scaled_ranks)
    observed_scaled = round(observed * 2.0)
    extreme = 0
    assignments = 1 << len(scaled_ranks)
    for mask in range(assignments):
        plus = sum(rank for index, rank in enumerate(scaled_ranks) if mask & (1 << index))
        if min(plus, total - plus) <= observed_scaled:
            extreme += 1
    p_value = extreme / assignments
    effect = (positive - negative) / (positive + negative)
    return observed, min(p_value, 1.0), effect, differences


def holm_adjust(p_values: Sequence[float]) -> list[float]:
    order = sorted(range(len(p_values)), key=lambda index: p_values[index])
    adjusted = [1.0] * len(p_values)
    running = 0.0
    count = len(p_values)
    for rank, index in enumerate(order):
        candidate = min((count - rank) * p_values[index], 1.0)
        running = max(running, candidate)
        adjusted[index] = running
    return adjusted


def bootstrap_median_ci(differences: Sequence[float], seed: int) -> tuple[float, float, float]:
    if not differences:
        return 0.0, 0.0, 0.0
    generator = random.Random(seed)
    estimates = []
    for _ in range(BOOTSTRAP_SAMPLES):
        sample = [differences[generator.randrange(len(differences))] for _ in differences]
        estimates.append(statistics.median(sample))
    estimates.sort()
    lower = estimates[round(0.025 * (len(estimates) - 1))]
    upper = estimates[round(0.975 * (len(estimates) - 1))]
    return statistics.median(differences), lower, upper


def cochran_q(matrix: Sequence[Sequence[bool]]) -> tuple[float, float]:
    if not matrix:
        return math.nan, math.nan
    method_count = len(matrix[0])
    column_sums = [sum(int(row[column]) for row in matrix) for column in range(method_count)]
    row_sums = [sum(int(value) for value in row) for row in matrix]
    total = sum(column_sums)
    denominator = method_count * total - sum(value * value for value in row_sums)
    if denominator == 0:
        return 0.0, 1.0
    statistic = (method_count - 1.0) * (method_count * sum(value * value for value in column_sums) - total * total) / denominator
    return statistic, chi_square_survival(statistic, method_count - 1)


def exact_mcnemar(baseline: Sequence[bool], optimized: Sequence[bool]) -> tuple[int, int, float]:
    baseline_only = sum(left and not right for left, right in zip(baseline, optimized))
    optimized_only = sum(not left and right for left, right in zip(baseline, optimized))
    discordant = baseline_only + optimized_only
    if discordant == 0:
        return baseline_only, optimized_only, 1.0
    tail = sum(math.comb(discordant, index) for index in range(0, min(baseline_only, optimized_only) + 1)) / (2 ** discordant)
    return baseline_only, optimized_only, min(2.0 * tail, 1.0)


def complete_by_task(trials: Iterable[Trial]) -> dict[str, dict[str, dict[str, Trial]]]:
    indexed: dict[str, dict[str, dict[str, Trial]]] = defaultdict(lambda: defaultdict(dict))
    for trial in trials:
        if trial.excluded:
            continue
        participant_methods = indexed[trial.task][trial.participant]
        if trial.method in participant_methods:
            raise ValueError(f"duplicate valid trial for {trial.participant}, {trial.task}, {trial.method}")
        participant_methods[trial.method] = trial
    return indexed


def metric_value(trial: Trial, metric: str) -> float:
    return {
        "time_s": trial.time_s,
        "errors": trial.errors,
        "navigation": trial.navigation,
        "readability": trial.readability,
        "cognitive_load": trial.cognitive_load,
    }[metric]


def descriptive_rows(trials: Sequence[Trial]) -> list[dict[str, object]]:
    valid = [trial for trial in trials if not trial.excluded]
    rows: list[dict[str, object]] = []
    for method in METHODS:
        for task in TASKS:
            group = [trial for trial in valid if trial.method == method and trial.task == task]
            row: dict[str, object] = {
                "method": method,
                "task": task,
                "n": len(group),
                "success_rate": sum(trial.success for trial in group) / len(group) if group else math.nan,
            }
            for label, values in {
                "time_s": [trial.time_s for trial in group],
                "errors": [trial.errors for trial in group],
                "navigation": [trial.navigation for trial in group],
                "readability": [trial.readability for trial in group],
                "cognitive_load": [trial.cognitive_load for trial in group],
            }.items():
                q1, median, q3 = quartiles(values)
                row[f"{label}_median"] = median
                row[f"{label}_q1"] = q1
                row[f"{label}_q3"] = q3
            rows.append(row)
    return rows


def inference_rows(trials: Sequence[Trial]) -> list[dict[str, object]]:
    indexed = complete_by_task(trials)
    output: list[dict[str, object]] = []
    for task_index, task in enumerate(TASKS):
        complete = {
            participant: methods
            for participant, methods in indexed.get(task, {}).items()
            if set(methods) == set(METHODS)
        }
        participants = sorted(complete)
        if len(participants) < MIN_COMPLETE_PARTICIPANTS:
            output.append({
                "task": task, "metric": "all", "test": "status", "comparison": "six conditions",
                "n": len(participants), "statistic": math.nan, "p_raw": math.nan, "p_holm": math.nan,
                "effect": math.nan, "median_difference": math.nan, "ci95_low": math.nan,
                "ci95_high": math.nan, "interpretation": "insufficient complete participants",
            })
            continue
        for metric_index, metric in enumerate(["time_s", "errors", "navigation", "readability", "cognitive_load"]):
            matrix = [[metric_value(complete[participant][method], metric) for method in METHODS] for participant in participants]
            statistic, p_value = friedman_test(matrix)
            output.append({
                "task": task, "metric": metric, "test": "Friedman", "comparison": "six conditions",
                "n": len(participants), "statistic": statistic, "p_raw": p_value, "p_holm": math.nan,
                "effect": math.nan, "median_difference": math.nan, "ci95_low": math.nan,
                "ci95_high": math.nan, "interpretation": "omnibus significant" if p_value < ALPHA else "no omnibus evidence",
            })
            if p_value >= ALPHA:
                continue
            baseline = [metric_value(complete[participant]["Baseline"], metric) for participant in participants]
            comparisons = []
            for method_index, method in enumerate(METHODS[1:], start=1):
                optimized = [metric_value(complete[participant][method], metric) for participant in participants]
                signed_rank, pair_p, effect, differences = exact_wilcoxon(baseline, optimized)
                median_difference, ci_low, ci_high = bootstrap_median_ci(
                    differences, seed=241_000 + task_index * 100 + metric_index * 10 + method_index,
                )
                comparisons.append({
                    "task": task, "metric": metric, "test": "exact Wilcoxon", "comparison": f"{method} - Baseline",
                    "n": len(differences), "statistic": signed_rank, "p_raw": pair_p, "effect": effect,
                    "median_difference": median_difference, "ci95_low": ci_low, "ci95_high": ci_high,
                })
            adjusted = holm_adjust([float(row["p_raw"]) for row in comparisons])
            for row, adjusted_p in zip(comparisons, adjusted):
                row["p_holm"] = adjusted_p
                row["interpretation"] = "Holm-significant" if adjusted_p < ALPHA else "not Holm-significant"
                output.append(row)
        success_matrix = [[complete[participant][method].success for method in METHODS] for participant in participants]
        q_statistic, q_p = cochran_q(success_matrix)
        output.append({
            "task": task, "metric": "success", "test": "Cochran Q", "comparison": "six conditions",
            "n": len(participants), "statistic": q_statistic, "p_raw": q_p, "p_holm": math.nan,
            "effect": math.nan, "median_difference": math.nan, "ci95_low": math.nan,
            "ci95_high": math.nan, "interpretation": "omnibus significant" if q_p < ALPHA else "no omnibus evidence",
        })
        if q_p < ALPHA:
            baseline_success = [complete[participant]["Baseline"].success for participant in participants]
            comparisons = []
            for method in METHODS[1:]:
                optimized_success = [complete[participant][method].success for participant in participants]
                baseline_only, optimized_only, pair_p = exact_mcnemar(baseline_success, optimized_success)
                comparisons.append({
                    "task": task, "metric": "success", "test": "exact McNemar", "comparison": f"{method} - Baseline",
                    "n": len(participants), "statistic": baseline_only + optimized_only, "p_raw": pair_p,
                    "effect": optimized_only - baseline_only, "median_difference": math.nan,
                    "ci95_low": math.nan, "ci95_high": math.nan,
                })
            adjusted = holm_adjust([float(row["p_raw"]) for row in comparisons])
            for row, adjusted_p in zip(comparisons, adjusted):
                row["p_holm"] = adjusted_p
                row["interpretation"] = "Holm-significant" if adjusted_p < ALPHA else "not Holm-significant"
                output.append(row)
    return output


def format_number(value: object, digits: int = 2) -> str:
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return f"{float(value):.{digits}f}"
    return ""


def write_csv(path: Path, rows: Sequence[dict[str, object]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_report(path: Path, trials: Sequence[Trial], planned_rows: int, descriptions: Sequence[dict[str, object]], inferences: Sequence[dict[str, object]]) -> None:
    valid = [trial for trial in trials if not trial.excluded]
    excluded = [trial for trial in trials if trial.excluded]
    lines = [
        "# 241-node behavior-tree developer usability analysis",
        "",
        "## Status",
        "",
    ]
    if not trials:
        lines.extend([
            f"Participant data have not been collected. The workbook contains {planned_rows} pre-generated trials and 0 completed observations.",
            "No human-usability result, significance test, or readability claim is reported.",
            "",
            "The study apparatus is ready for 12 participants; after collection, rerun this script to generate descriptive and inferential outputs.",
        ])
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return
    lines.extend([
        f"Completed rows: {len(trials)} / {planned_rows}; valid rows: {len(valid)}; excluded rows: {len(excluded)}.",
        "",
        "## Descriptive results",
        "",
        "| Method | Task | N | Success | Time Median [Q1, Q3] | Navigation Median [Q1, Q3] | Readability | Cognitive load |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: |",
    ])
    for row in descriptions:
        if not row["n"]:
            continue
        lines.append(
            f"| {row['method']} | {row['task']} | {row['n']} | {float(row['success_rate']):.1%} | "
            f"{format_number(row['time_s_median'])} [{format_number(row['time_s_q1'])}, {format_number(row['time_s_q3'])}] | "
            f"{format_number(row['navigation_median'])} [{format_number(row['navigation_q1'])}, {format_number(row['navigation_q3'])}] | "
            f"{format_number(row['readability_median'])} | {format_number(row['cognitive_load_median'])} |"
        )
    lines.extend([
        "",
        "## Inferential status",
        "",
        "The CSV output contains the pre-registered Friedman/Cochran omnibus tests and, only after a significant omnibus result, Holm-corrected pairwise tests. A negative optimized-minus-baseline difference favours the optimization for time, errors, navigation and cognitive load; a positive difference favours it for readability.",
        "",
    ])
    insufficient = [row for row in inferences if row["test"] == "status"]
    if insufficient:
        lines.append("At least one task has fewer than eight complete participant sets; inferential conclusions are therefore withheld and results are descriptive only.")
    else:
        lines.append("All tasks meet the pre-registered minimum of eight complete participant sets. Interpret individual comparisons only when their Holm-adjusted p-value is below 0.05 and success has not worsened.")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_self_test() -> None:
    matrix = [[10 + participant, 8 + participant, 9 + participant] for participant in range(12)]
    statistic, p_value = friedman_test(matrix)
    assert statistic > 0.0 and p_value < 0.05
    baseline = [20.0 + participant for participant in range(12)]
    optimized = [value - 5.0 for value in baseline]
    _, wilcoxon_p, effect, differences = exact_wilcoxon(baseline, optimized)
    assert wilcoxon_p < 0.05 and effect == -1.0 and statistics.median(differences) == -5.0
    adjusted = holm_adjust([0.01, 0.02, 0.20, 0.04, 0.50])
    assert all(0.0 <= value <= 1.0 for value in adjusted)
    q_statistic, q_p = cochran_q([[True, True, False], [False, True, False], [True, True, True]])
    assert q_statistic >= 0.0 and 0.0 <= q_p <= 1.0
    _, _, mcnemar_p = exact_mcnemar([True] * 10, [False] * 10)
    assert mcnemar_p < 0.05
    print("HUMAN_STUDY_ANALYSIS_SELF_TEST passed=5 failed=0 (synthetic values not written)")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=Path("research/display_optimization/behavior_tree_human_comparison_study.xlsx"))
    parser.add_argument("--output-dir", type=Path, default=Path("research/display_optimization/human_study_results"))
    parser.add_argument("--self-test", action="store_true")
    arguments = parser.parse_args()
    if arguments.self_test:
        run_self_test()
        return 0
    trials, planned_rows = load_trials(arguments.workbook)
    descriptions = descriptive_rows(trials)
    inferences = inference_rows(trials) if trials else []
    arguments.output_dir.mkdir(parents=True, exist_ok=True)
    descriptive_fields = [
        "method", "task", "n", "success_rate",
        "time_s_median", "time_s_q1", "time_s_q3",
        "errors_median", "errors_q1", "errors_q3",
        "navigation_median", "navigation_q1", "navigation_q3",
        "readability_median", "readability_q1", "readability_q3",
        "cognitive_load_median", "cognitive_load_q1", "cognitive_load_q3",
    ]
    inference_fields = [
        "task", "metric", "test", "comparison", "n", "statistic", "p_raw", "p_holm",
        "effect", "median_difference", "ci95_low", "ci95_high", "interpretation",
    ]
    write_csv(arguments.output_dir / "descriptive_summary.csv", descriptions, descriptive_fields)
    write_csv(arguments.output_dir / "inferential_tests.csv", inferences, inference_fields)
    write_report(arguments.output_dir / "analysis_report.md", trials, planned_rows, descriptions, inferences)
    print(
        "HUMAN_STUDY_ANALYSIS completed=%d planned=%d valid=%d excluded=%d output=%s"
        % (len(trials), planned_rows, sum(not trial.excluded for trial in trials), sum(trial.excluded for trial in trials), arguments.output_dir)
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
